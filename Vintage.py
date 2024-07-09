#!/usr/bin/env python
# coding: utf-8

# # For Total Portfolio

# In[ ]:


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def calculate_percentage(excel_file):
    # Load the Excel file into a dictionary of DataFrames
    xls = pd.ExcelFile(excel_file)
    sheet_data = {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}

    percentages_all_sheets = []

    for idx, last_sheet_name in enumerate(sheet_data.keys()):
        last_sheet_data = sheet_data[last_sheet_name]

        # Step 1: Calculate the sum of "Disbursed amount in AZN" for the current last sheet
        sum_disbursed_amount = last_sheet_data['Disbursed amount'].sum()

        percentages = []

        # Step 2 and 3: Calculate percentages using vectorized calculations
        for sheet_name in list(sheet_data.keys())[:idx + 1]:
            sheet = sheet_data[sheet_name]
            sum_outstanding_overdue = sheet.loc[
                (sheet['Disbursed month'] == last_sheet_name) &
                (sheet['Overdue bucket > 30 days'] == 'PAR > 30 Days'),
                'Outstanding amount'
            ].sum()

            percentage = (sum_outstanding_overdue / sum_disbursed_amount) * 100
            percentages.append(round(percentage, 1))

        percentages.reverse()
        percentages_all_sheets.append(percentages)

    # Create a DataFrame using the calculated percentages
    df = pd.DataFrame(percentages_all_sheets, columns=[f"{i}" for i in range(1, len(sheet_data) + 1)])
    df.index = sheet_data.keys()

    # Reverse the order of rows in the DataFrame
    df = df.iloc[::-1]

    # Write the DataFrame to a new Excel file with styles
    output_excel_file = 'Final_Vintage.xlsx'
    writer = pd.ExcelWriter(output_excel_file, engine='openpyxl')
    writer.book = Workbook()
    df.to_excel(writer, index=True, sheet_name='Sheet1')

    # Apply styles to the entire DataFrame
    sheet = writer.sheets['Sheet1']
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(sheet_data) + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12, bold=True)

    writer.save()

    return df


# Replace 'your_excel_file.xlsx' with the path to your actual Excel file
if __name__ == "__main__":
    excel_file_path = r"C:\Users\username\Desktop\filename.xlsx"
    result_df = calculate_percentage(excel_file_path)
    print(result_df)


# # Based on Different Products

# In[ ]:


import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def calculate_percentage(excel_file, product_name):
    # Load the Excel file into a dictionary of DataFrames
    xls = pd.ExcelFile(excel_file)
    sheet_data = {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}

    percentages_all_sheets = []

    for idx, last_sheet_name in enumerate(sheet_data.keys()):
        last_sheet_data = sheet_data[last_sheet_name]

        # Step 1: Calculate the sum of "Disbursed amount" for the current last sheet
        sum_disbursed_amount = last_sheet_data.loc[
            last_sheet_data['Product'] == product_name, 
            'Disbursed amount'
        ].sum()

        percentages = []

        # Step 2 and 3: Calculate percentages using vectorized calculations
        for sheet_name in list(sheet_data.keys())[:idx + 1]:
            sheet = sheet_data[sheet_name]

            sum_outstanding_overdue = sheet.loc[
                (sheet['Disbursed month'] == last_sheet_name) &
                (sheet['Overdue bucket > 30 days'] == 'PAR > 30 Days') &
                (sheet['Product'] == product_name),
                'Outstanding amount'
            ].sum()

            percentage = (sum_outstanding_overdue / sum_disbursed_amount) * 100
            percentages.append(percentage)

        percentages.reverse()
        percentages_all_sheets.append(percentages)

    # Create a DataFrame using the calculated percentages
    df = pd.DataFrame(percentages_all_sheets, columns=[f"{i}" for i in range(1, len(sheet_data) + 1)])
    df.index = sheet_data.keys()

    # Reverse the order of rows in the DataFrame
    df = df.iloc[::-1]

    # Write the DataFrame to a new Excel file with styles
    output_excel_file = 'Final_Vintage_Product.xlsx'
    writer = pd.ExcelWriter(output_excel_file, engine='openpyxl')
    writer.book = Workbook()
    df.to_excel(writer, index=True, sheet_name='Sheet1')

    # Apply styles to the entire DataFrame
    sheet = writer.sheets['Sheet1']
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(sheet_data) + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=12, bold=True)

    writer.save()

    return df


if __name__ == "__main__":
    # Replace 'ProductName' with the actual product name you need
    product_name = 'ProductName'
    # Replace 'your_excel_file.xlsx' with the path to your actual Excel file
    excel_file_path = r"C:\Users\username\Desktop\filename.xlsx"
    result_df = calculate_percentage(excel_file_path, product_name)
    print(result_df)

